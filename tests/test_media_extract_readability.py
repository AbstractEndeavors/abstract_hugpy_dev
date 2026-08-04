"""Backend tests for /ml/extract's READABILITY posture (operator ask 2026-08-04, k64:
"attachments often aren't recognized/readable and the failures are mushy").

Covers (GPU-free, script-style with a __main__ guard like the sibling tests):

  1. .xlsx happy path — a workbook built here with openpyxl comes back ok, with a
     per-sheet `pages` list and tab-separated cell text (formula cells read as
     their cached VALUE, not "=A1+B1").
  2. Legacy binary .doc / .xls decline with a SPECIFIC, actionable error naming
     the format and the one-step fix — not the old generic "isn't readable as
     text" that made every failure look the same.
  3. A scanned/no-text-layer PDF says exactly that (OCR isn't wired up), rather
     than the old shared "empty or image-only/scanned" hedge.
  4. EVERY decline path emits a `logger.warning("media extract declined: …")`,
     so a file the chat couldn't read leaves a trace in the server log.
  5. PARTIAL PDFs degrade instead of failing (operator ask 2026-08-04, k65): a
     page that throws is recorded + skipped, the run keeps going, and the result
     carries pages_total / pages_extracted / warnings ("page 4 failed: …") plus
     cheap document meta (title/author/page_count). A PDF that yields NOTHING
     still takes the k64 decline path, unchanged.

Isolation: media_extract.UPLOADS_HOME is repointed to a PRIVATE temp dir (the
_selftest idiom the sibling tests use for media_bus.DB_PATH) so the real storage
root is never written to, and the path jail is still exercised for real.

Run:
  abstract_hugpy_dev/venv/bin/python -m pytest tests/test_media_extract_readability.py -q
  abstract_hugpy_dev/venv/bin/python tests/test_media_extract_readability.py
"""
import logging
import os
import sys
import tempfile
from contextlib import contextmanager
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

_MODULE = "abstract_hugpy_dev.flask_app.app.functions.media_extract"


def _private_root():
    """Import media_extract with its storage jail repointed at a temp dir."""
    from abstract_hugpy_dev.flask_app.app.functions import media_extract
    tmpdir = tempfile.mkdtemp(prefix="hugpy_test_extract_")
    media_extract.UPLOADS_HOME = tmpdir
    media_extract.DEFAULT_ROOT = tmpdir
    return media_extract, tmpdir


@contextmanager
def _captured_warnings():
    """Collect formatted WARNING+ records from the module logger.

    A plain handler (not caplog) so the __main__ runner works too; the caplog
    test below asserts the same thing through pytest's fixture.
    """
    records: list[str] = []

    class _Sink(logging.Handler):
        def emit(self, record):
            records.append(record.getMessage())

    logger = logging.getLogger(_MODULE)
    sink = _Sink(level=logging.WARNING)
    prior_level = logger.level
    logger.addHandler(sink)
    logger.setLevel(logging.WARNING)
    try:
        yield records
    finally:
        logger.removeHandler(sink)
        logger.setLevel(prior_level)


# --------------------------------------------------------------------------- #
# 1) .xlsx is genuinely READ (was declined by the binary sniff before k64).
# --------------------------------------------------------------------------- #
def test_xlsx_happy_path():
    media_extract, root = _private_root()
    import openpyxl

    path = os.path.join(root, "quarter.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws.append(["Region", "Q1", "Q2"])
    ws.append(["North", 1200, 1350])
    ws.append([None, None, None])          # blank row -> dropped, not a blank line
    ws.append(["South", 900, 1100])
    ws2 = wb.create_sheet("Notes")
    ws2.append(["renewals slipped to Q3"])
    wb.save(path)

    res = media_extract.extract_document(path)
    assert res["ok"] is True, res
    assert res["kind"] == "xlsx", res
    text = res["text"]
    assert "Region\tQ1\tQ2" in text, text          # tab-separated: column structure survives
    assert "North\t1200\t1350" in text, text
    assert "South\t900\t1100" in text, text
    assert "# Sheet: Revenue" in text and "# Sheet: Notes" in text, text
    assert "renewals slipped to Q3" in text, text
    assert "\n\n\n" not in text, "the wholly-empty row should not survive as blank lines"
    # pages == one entry per non-empty sheet (mirrors the PDF per-page shape).
    names = [p["name"] for p in res["pages"]]
    assert names == ["Revenue", "Notes"], res["pages"]
    assert res["chars"] == len(text), res
    print("[1] PASS  .xlsx reads to per-sheet tab-separated text")


def test_xlsx_reads_cached_formula_values():
    """data_only=True: a formula cell must come back as its VALUE, not '=SUM(...)'."""
    media_extract, root = _private_root()
    import openpyxl

    path = os.path.join(root, "formula.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = 2, 3
    ws["C1"] = "=A1+B1"
    wb.save(path)

    res = media_extract.extract_document(path)
    # openpyxl writes no cached value for a formula it never evaluated, so the
    # cell reads as empty — the contract that matters is that the raw formula
    # STRING never leaks into the model's context.
    assert "=A1+B1" not in res.get("text", ""), res
    print("[2] PASS  formula cells never leak '=…' into the extracted text")


# --------------------------------------------------------------------------- #
# 2) Legacy binary formats decline with a SPECIFIC, actionable error.
# --------------------------------------------------------------------------- #
def test_legacy_doc_specific_error():
    media_extract, root = _private_root()

    doc = os.path.join(root, "contract.doc")
    with open(doc, "wb") as fh:
        fh.write(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)  # OLE2 header
    res = media_extract.extract_document(doc)
    assert res["ok"] is False, res
    assert res["kind"] == "unsupported", res
    err = res["error"]
    assert ".doc" in err and ".docx" in err, err   # names the format AND the fix
    assert "isn't readable as text" not in err, "must not fall back to the old generic message"
    assert res["name"] == "contract.doc", res

    xls = os.path.join(root, "books.xls")
    with open(xls, "wb") as fh:
        fh.write(b"\xd0\xcf\x11\xe0" + b"\x00" * 32)
    res_xls = media_extract.extract_document(xls)
    assert res_xls["ok"] is False and ".xlsx" in res_xls["error"], res_xls
    print("[3] PASS  legacy .doc/.xls decline by NAME with the conversion that fixes it")


# --------------------------------------------------------------------------- #
# 3) A no-text-layer PDF says "scanned; no OCR", not a generic hedge.
# --------------------------------------------------------------------------- #
def test_scanned_pdf_error_is_specific():
    media_extract, root = _private_root()

    # A structurally valid one-page PDF with NO text operators — exactly what a
    # scan looks like to a text extractor.
    path = os.path.join(root, "scan.pdf")
    body = (
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 200 200]>>endobj\n"
        b"trailer<</Root 1 0 R>>\n%%EOF\n"
    )
    with open(path, "wb") as fh:
        fh.write(body)

    res = media_extract.extract_document(path)
    assert res["ok"] is False, res
    assert res["kind"] == "pdf", res
    err = res["error"].lower()
    assert "ocr" in err, err
    assert ("no text layer" in err or "corrupt" in err or "password" in err), err
    print("[4] PASS  a text-layer-free PDF explains itself (scan / OCR not wired up)")


# --------------------------------------------------------------------------- #
# 4) EVERY decline logs a warning (they were invisible in ops before k64).
# --------------------------------------------------------------------------- #
def _decline_cases(root: str) -> list[str]:
    """One file per decline branch: unsupported, unknown-binary, empty text."""
    legacy = os.path.join(root, "old.doc")
    with open(legacy, "wb") as fh:
        fh.write(b"\xd0\xcf\x11\xe0")
    binary = os.path.join(root, "blob.bin")
    with open(binary, "wb") as fh:
        fh.write(b"\x00\x01\x02\x03" * 32)
    empty = os.path.join(root, "empty.txt")
    with open(empty, "w", encoding="utf-8") as fh:
        fh.write("   \n")
    return [legacy, binary, empty]


def test_every_decline_logs_a_warning():
    media_extract, root = _private_root()
    paths = _decline_cases(root)

    with _captured_warnings() as records:
        for p in paths:
            res = media_extract.extract_document(p)
            assert res["ok"] is False, (p, res)

    assert len(records) == len(paths), records
    for msg, p in zip(records, paths):
        assert msg.startswith("media extract declined: "), msg
        assert os.path.basename(p) in msg, (msg, p)
        assert msg.rstrip().endswith(")"), "the reason must be logged alongside the name"
    print("[5] PASS  every decline branch logs 'media extract declined: <name> (<why>)'")


def test_decline_logs_via_caplog(caplog):
    """The same guarantee through pytest's caplog (the ops-visible path)."""
    media_extract, root = _private_root()
    path = os.path.join(root, "legacy.xls")
    with open(path, "wb") as fh:
        fh.write(b"\xd0\xcf\x11\xe0")

    with caplog.at_level(logging.WARNING, logger=_MODULE):
        res = media_extract.extract_document(path)
    assert res["ok"] is False, res
    assert "media extract declined: legacy.xls" in caplog.text, caplog.text
    assert ".xlsx" in caplog.text, caplog.text
    print("[6] PASS  the decline warning is visible through caplog")


# A successful read must NOT log a decline (no crying wolf in the ops log).
def test_success_logs_nothing():
    media_extract, root = _private_root()
    path = os.path.join(root, "readme.md")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("# hello\nsome content\n")
    with _captured_warnings() as records:
        res = media_extract.extract_document(path)
    assert res["ok"] is True, res
    assert records == [], records
    print("[7] PASS  a successful read logs no decline")


# --------------------------------------------------------------------------- #
# 5) A partly-broken PDF degrades: partial text + per-page warnings + counts.
#    (operator ask 2026-08-04, k65)
# --------------------------------------------------------------------------- #
def _build_pdf(page_texts, title=None, author=None) -> bytes:
    """A real, parseable multi-page PDF — one text operator per non-empty page.

    Hand-assembled (with a correct xref) rather than written by a library: neither
    reportlab nor pypdf's writer is in the venv, and PyPDF2's writer can only add
    BLANK pages, which is exactly what this test must NOT have everywhere.
    """
    n = len(page_texts)
    page_ids = list(range(4, 4 + 2 * n, 2))
    content_ids = [i + 1 for i in page_ids]
    info_id = (4 + 2 * n) if (title or author) else None

    objs = {
        1: b"<< /Type /Catalog /Pages 2 0 R >>",
        2: (b"<< /Type /Pages /Kids [" + b" ".join(b"%d 0 R" % i for i in page_ids)
            + b"] /Count %d >>" % n),
        3: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    }
    for pid, cid, txt in zip(page_ids, content_ids, page_texts):
        objs[pid] = (b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 300 300] "
                     b"/Resources << /Font << /F1 3 0 R >> >> /Contents %d 0 R >>" % cid)
        stream = (b"BT /F1 14 Tf 40 200 Td (" + txt.encode("ascii") + b") Tj ET\n") if txt else b""
        objs[cid] = b"<< /Length %d >>\nstream\n" % len(stream) + stream + b"endstream"
    if info_id:
        info = []
        if title:
            info.append(b"/Title (" + title.encode("ascii") + b")")
        if author:
            info.append(b"/Author (" + author.encode("ascii") + b")")
        objs[info_id] = b"<< " + b" ".join(info) + b" >>"

    out, offsets = bytearray(b"%PDF-1.4\n"), {}
    for oid in sorted(objs):
        offsets[oid] = len(out)
        out += b"%d 0 obj\n" % oid + objs[oid] + b"\nendobj\n"
    xref_at, last = len(out), max(objs)
    out += b"xref\n0 %d\n" % (last + 1) + b"0000000000 65535 f \n"
    for oid in range(1, last + 1):
        out += (b"%010d 00000 n \n" % offsets[oid]) if oid in offsets else b"0000000000 65535 f \n"
    out += b"trailer\n<< /Size %d /Root 1 0 R" % (last + 1)
    if info_id:
        out += b" /Info %d 0 R" % info_id
    out += b" >>\nstartxref\n%d\n%%%%EOF\n" % xref_at
    return bytes(out)


@contextmanager
def _engine(name: str, replacement):
    """Swap a parser module (pdfplumber / PyPDF2) for the duration of a call.

    media_extract imports both LAZILY inside the extractor, so a stand-in in
    sys.modules is what the code under test picks up. Used to make a page throw
    on demand: the unit under test is OUR per-page tolerance, not the parser's
    taste in malformed files.
    """
    prior = sys.modules.get(name)
    sys.modules[name] = replacement
    try:
        yield
    finally:
        if prior is None:
            sys.modules.pop(name, None)
        else:
            sys.modules[name] = prior


class _FakePage:
    def __init__(self, text=None, boom=None):
        self._text, self._boom = text, boom

    def extract_text(self):
        if self._boom:
            raise ValueError(self._boom)
        return self._text


class _FakePdf:
    def __init__(self, pages, metadata):
        self.pages, self.metadata = pages, metadata

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


class _FakePdfplumber:
    """Stand-in module object: `import pdfplumber` just binds sys.modules[name]."""

    def __init__(self, pages, metadata=None, open_error=None):
        self._pages, self._metadata, self._open_error = pages, metadata or {}, open_error

    def open(self, path):
        if self._open_error:
            raise RuntimeError(self._open_error)
        return _FakePdf(self._pages, self._metadata)


def test_pdf_partial_pages_degrade_with_warnings():
    media_extract, root = _private_root()

    path = os.path.join(root, "partial.pdf")
    with open(path, "wb") as fh:
        fh.write(_build_pdf(["placeholder body: the fake engine answers"]))

    fake = _FakePdfplumber(
        pages=[
            _FakePage(text="page one body"),
            _FakePage(boom="stream is damaged"),
            _FakePage(text="page three body"),
            _FakePage(boom="unexpected end of object"),
            _FakePage(text="page five body"),
        ],
        metadata={"Title": "Field Notes", "Author": "Ada Lovelace"},
    )
    with _engine("pdfplumber", fake):
        res = media_extract.extract_document(path)

    assert res["ok"] is True, res                      # 2 bad pages do NOT sink the read
    assert "page one body" in res["text"], res
    assert "page three body" in res["text"], res
    assert "page five body" in res["text"], res
    assert res["pages_total"] == 5, res
    assert res["pages_extracted"] == 3, res
    warnings = res["warnings"]
    assert len(warnings) == 2, warnings
    assert warnings[0] == "page 2 failed: stream is damaged", warnings
    assert warnings[1].startswith("page 4 failed: "), warnings
    # The failed pages are still IN `pages`, carrying their reason (not dropped).
    assert res["pages"][1] == {"index": 1, "text": "", "error": "stream is damaged"}, res["pages"]
    assert res["meta"] == {"page_count": 5, "title": "Field Notes", "author": "Ada Lovelace"}, res
    print("[8] PASS  a partly-broken PDF yields partial text + per-page warnings + counts")


def test_pdf_counts_and_metadata_on_a_real_pdf():
    """Real parsers, real file: counts + Info-dict metadata, no warnings when clean."""
    media_extract, root = _private_root()

    path = os.path.join(root, "report.pdf")
    with open(path, "wb") as fh:
        fh.write(_build_pdf(
            ["Revenue rose in Q1", "", "Renewals slipped to Q3"],
            title="Quarterly Report", author="Ada",
        ))

    res = media_extract.extract_document(path)
    assert res["ok"] is True, res
    assert res["kind"] == "pdf", res
    assert "Revenue rose in Q1" in res["text"] and "Renewals slipped to Q3" in res["text"], res
    assert res["pages_total"] == 3, res
    assert res["pages_extracted"] == 2, res            # the blank page yields nothing
    assert res["warnings"] == [], res                  # blank != failed: no false alarm
    assert res["meta"]["page_count"] == 3, res
    assert res["meta"]["title"] == "Quarterly Report", res
    assert res["meta"]["author"] == "Ada", res
    print("[9] PASS  a real PDF reports pages_total/pages_extracted + title/author/page_count")


def test_pdf_fallback_engine_is_page_tolerant():
    """Engine 2 (PyPDF2) is fault-tolerant too — with a genuinely broken page.

    The page's /Contents points at a missing object, which makes PyPDF2 raise on
    that page alone. pdfplumber is stubbed OUT (open fails) to force the fallback.
    """
    media_extract, root = _private_root()

    data = _build_pdf(["First page words", "Second page words", "Third page words"])
    data = data.replace(b"/Contents 7 0 R", b"/Contents 99 0 R")   # page 2 -> missing object
    path = os.path.join(root, "fallback.pdf")
    with open(path, "wb") as fh:
        fh.write(data)

    with _engine("pdfplumber", _FakePdfplumber(pages=[], open_error="no pdfplumber here")):
        res = media_extract.extract_document(path)

    assert res["ok"] is True, res
    assert "First page words" in res["text"] and "Third page words" in res["text"], res
    assert res["pages_total"] == 3, res
    assert res["pages_extracted"] == 2, res
    assert len(res["warnings"]) == 1 and res["warnings"][0].startswith("page 2 failed: "), res
    print("[10] PASS  the PyPDF2 fallback skips the one broken page and keeps the rest")


def test_fully_broken_pdf_still_declines():
    """Zero extracted pages -> the k64 decline path, wording unchanged."""
    media_extract, root = _private_root()

    path = os.path.join(root, "shredded.pdf")
    with open(path, "wb") as fh:
        fh.write(b"%PDF-1.4\n" + b"\x00\xff\x13not a pdf at all\x00" * 40)

    with _captured_warnings() as records:
        res = media_extract.extract_document(path)
    assert res["ok"] is False, res
    assert res["kind"] == "pdf", res
    err = res["error"].lower()
    assert "corrupt" in err or "password" in err, err   # nothing enumerated -> not "scanned"
    assert res["text"] == "", res
    assert records and records[0].startswith("media extract declined: shredded.pdf"), records
    print("[11] PASS  a wholly unreadable PDF still declines (k64 path untouched)")


class _FakeCaplog:
    """Minimal caplog stand-in so the __main__ runner covers the caplog test too."""

    def __init__(self):
        self.text = ""

    @contextmanager
    def at_level(self, level, logger=None):
        with _captured_warnings() as records:
            yield
        self.text = "\n".join(records)


def _run_all():
    test_xlsx_happy_path()
    test_xlsx_reads_cached_formula_values()
    test_legacy_doc_specific_error()
    test_scanned_pdf_error_is_specific()
    test_every_decline_logs_a_warning()
    test_decline_logs_via_caplog(_FakeCaplog())
    test_success_logs_nothing()
    test_pdf_partial_pages_degrade_with_warnings()
    test_pdf_counts_and_metadata_on_a_real_pdf()
    test_pdf_fallback_engine_is_page_tolerant()
    test_fully_broken_pdf_still_declines()
    print("\nALL media-extract readability backend checks passed")


if __name__ == "__main__":
    _run_all()
